"""Upload a 4-week roster-request grid and load it into the database.

The spreadsheet is a *wide grid*, one row per person::

    | name          | role      | fte | 2026-08-03 | 2026-08-04 | ... |
    | Dr Alice Chen | REGISTRAR | 1.0 | D          | x          | ... |
    | Dr Bob Singh  | RESIDENT  | 0.5 | N          |            | ... |

The first three columns are stable each period (they describe the staff member).
Every column after that is a single date, and each cell is a request code:

    D -> wants a DAY shift      N -> wants a NIGHT shift      x -> wants OFF

A blank cell means "no request" and is skipped.

Because the `request` table stores one row per (staff, date), the grid has to be
*unpivoted*: a person with 28 date columns becomes up to 28 request rows.
"""

import datetime
import io

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import delete
from sqlalchemy.orm import Session

from db.database import get_db
from db.models import Request, Staff

router = APIRouter()

FIXED_HEADERS = ["name", "role", "fte"]
VALID_ROLES = {"REGISTRAR", "RESIDENT"}
PERIOD_LENGTH_DAYS = 28  # 4 weeks

# Cell code -> what we store in request.request_type. Change the values here if
# your solver expects different shift names (this is the one place to edit).
CODE_TO_REQUEST_TYPE = {
    "D": "DAY",
    "N": "NIGHT",
    "X": "OFF",
}


def _column_letter(index: int) -> str:
    """0-based column index -> Excel letter (0 -> 'A', 26 -> 'AA'). For errors."""
    letters = ""
    index += 1
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


@router.post("/upload/roster-requests")
async def upload_roster_requests(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Accept a 4-week request grid (.xlsx) and load it into staff + request."""
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "Please upload an .xlsx file")

    contents = await file.read()
    try:
        # data_only reads computed values, not formulas; read_only is faster.
        workbook = load_workbook(io.BytesIO(contents), data_only=True, read_only=True)
    except Exception as exc:
        raise HTTPException(400, f"Could not read that spreadsheet: {exc}") from exc

    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)

    try:
        header = next(rows)
    except StopIteration:
        raise HTTPException(400, "The spreadsheet is empty") from None

    # --- Validate the header ------------------------------------------------
    if header is None or len(header) < len(FIXED_HEADERS) + 1:
        raise HTTPException(
            400,
            "Expected columns: name, role, fte, then one column per date.",
        )

    fixed = [str(h).strip().lower() if h is not None else "" for h in header[:3]]
    if fixed != FIXED_HEADERS:
        raise HTTPException(
            400,
            f"First three columns must be name, role, fte — got {fixed}.",
        )

    # Every remaining header cell must be a date.
    date_columns: list[tuple[int, datetime.date]] = []
    bad_headers: list[str] = []
    for col_index, value in enumerate(header[3:], start=3):
        parsed = _coerce_date(value)
        if parsed is None:
            bad_headers.append(f"{_column_letter(col_index)}='{value}'")
        else:
            date_columns.append((col_index, parsed))

    if bad_headers:
        raise HTTPException(
            400,
            "These column headers are not dates: " + ", ".join(bad_headers),
        )

    warnings: list[str] = []
    if len(date_columns) != PERIOD_LENGTH_DAYS:
        warnings.append(
            f"Expected {PERIOD_LENGTH_DAYS} date columns (4 weeks); "
            f"found {len(date_columns)}."
        )

    # --- Validate every row, collecting errors before writing anything ------
    parsed_rows: list[dict] = []
    errors: list[str] = []

    for row_number, row in enumerate(rows, start=2):  # row 2 = first data row
        if row is None or all(cell is None for cell in row):
            continue  # skip fully-blank rows

        name = str(row[0]).strip() if row[0] is not None else ""
        if not name:
            continue  # a row with codes but no name is treated as blank

        role = str(row[1]).strip().upper() if row[1] is not None else ""
        if role not in VALID_ROLES:
            errors.append(
                f"Row {row_number} ({name}): role must be REGISTRAR or RESIDENT, "
                f"got '{role or '(blank)'}'."
            )
            continue

        fte = _coerce_fte(row[2] if len(row) > 2 else None)
        if fte is None:
            raw_fte = row[2] if len(row) > 2 else None
            errors.append(
                f"Row {row_number} ({name}): FTE '{raw_fte}' is not a number "
                "between 0 and 1."
            )
            continue

        requests: dict[datetime.date, str] = {}
        for col_index, date in date_columns:
            raw = row[col_index] if col_index < len(row) else None
            if raw is None or str(raw).strip() == "":
                continue  # no request for this day
            code = str(raw).strip().upper()
            if code not in CODE_TO_REQUEST_TYPE:
                errors.append(
                    f"Row {row_number} ({name}), {date} "
                    f"[cell {_column_letter(col_index)}{row_number}]: "
                    f"unknown code '{raw}' — use "
                    f"{', '.join(sorted(CODE_TO_REQUEST_TYPE))} or leave blank."
                )
                continue
            requests[date] = CODE_TO_REQUEST_TYPE[code]

        parsed_rows.append(
            {"name": name, "role": role, "fte": fte, "requests": requests}
        )

    if errors:
        # Reject the whole file rather than half-importing it.
        raise HTTPException(400, {"message": "File not imported", "errors": errors})

    if not parsed_rows:
        raise HTTPException(400, "No data rows found in the spreadsheet.")

    # --- Write: upsert staff, then replace their requests for these dates ---
    existing_staff = {s.name: s for s in db.query(Staff).all()}
    staff_ids: list[int] = []

    for entry in parsed_rows:
        staff = existing_staff.get(entry["name"])
        if staff is None:
            staff = Staff(name=entry["name"])
            db.add(staff)
            existing_staff[entry["name"]] = staff
        staff.classification = entry["role"]
        staff.employment_fraction = entry["fte"]
        db.flush()  # assign staff_id for newly-created rows
        entry["staff_id"] = staff.staff_id
        staff_ids.append(staff.staff_id)

    all_dates = [date for _, date in date_columns]
    # Make re-uploads idempotent: clear existing requests in this window first.
    db.execute(
        delete(Request).where(
            Request.staff_id.in_(staff_ids),
            Request.date.in_(all_dates),
        )
    )

    new_requests = [
        Request(
            staff_id=entry["staff_id"],
            date=date,
            request_type=request_type,
            approved=None,  # unreviewed
        )
        for entry in parsed_rows
        for date, request_type in entry["requests"].items()
    ]
    db.add_all(new_requests)
    db.commit()

    return {
        "filename": file.filename,
        "staff_upserted": len(parsed_rows),
        "requests_imported": len(new_requests),
        "period": {
            "start": min(all_dates).isoformat(),
            "end": max(all_dates).isoformat(),
            "days": len(all_dates),
        },
        "warnings": warnings,
    }


@router.get("/upload/roster-requests/template")
def download_template(start_date: datetime.date | None = None):
    """Return a blank .xlsx laid out exactly the way the upload expects.

    ?start_date=YYYY-MM-DD sets the first date column (defaults to next Monday).
    Hand this to staff so their columns always line up with the parser.
    """
    if start_date is None:
        today = datetime.date.today()
        start_date = today + datetime.timedelta(days=(7 - today.weekday()) % 7 or 7)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "requests"

    dates = [start_date + datetime.timedelta(days=i) for i in range(PERIOD_LENGTH_DAYS)]
    sheet.append([*FIXED_HEADERS, *dates])
    for cell in sheet[1][len(FIXED_HEADERS):]:
        cell.number_format = "yyyy-mm-dd"

    # A couple of illustrative rows people can overwrite or delete.
    sheet.append(["Dr Example One", "REGISTRAR", 1.0])
    sheet.append(["Dr Example Two", "RESIDENT", 0.5])

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"roster-requests-{start_date.isoformat()}.xlsx"
    return StreamingResponse(
        buffer,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _coerce_date(value) -> datetime.date | None:
    """Turn a header cell into a date, whether Excel stored it as a date or text."""
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    if value is None:
        return None
    try:
        return datetime.date.fromisoformat(str(value).strip()[:10])
    except ValueError:
        return None


def _coerce_fte(value) -> float | None:
    """Validate FTE is a number in (0, 1]."""
    try:
        fte = float(value)
    except (ValueError, TypeError):
        return None
    if 0 < fte <= 1:
        return fte
    return None
